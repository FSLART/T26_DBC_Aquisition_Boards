import os
import cantools
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Styles
header_font = Font(bold=True)
header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
message_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def auto_adjust_column_width(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

def write_message_table(ws, message, start_row):
    # Table header
    headers = [
        "Signal Name", "Start Bit", "Length (bits)", "Byte Order", "Signed",
        "Factor", "Offset", "Min", "Max", "Unit", "Choices"
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Table rows
    for row_index, sig in enumerate(message.signals, start=start_row + 1):
        row_data = [
            sig.name,
            sig.start,
            sig.length,
            "Intel" if sig.byte_order == "little_endian" else "Motorola",
            sig.is_signed,
            sig.scale,
            sig.offset,
            sig.minimum,
            sig.maximum,
            sig.unit,
            ", ".join(f"{k}={v}" for k, v in sig.choices.items()) if sig.choices else ""
        ]
        for col_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = thin_border

    return row_index + 2  # leave a blank row between tables

def export_dbc_to_excel(dbc_files, output_filename="dbc_signals.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    for dbc_file in dbc_files:
        try:
            db = cantools.database.load_file(dbc_file)
        except Exception as e:
            print(f"Error loading {dbc_file}: {e}")
            continue

        sheet_name = os.path.splitext(os.path.basename(dbc_file))[0][:31]
        ws = wb.create_sheet(title=sheet_name)

        current_row = 1
        for msg in db.messages:
            senders = ", ".join(msg.senders) if msg.senders else "Unknown"
            label_cell_1 = ws.cell(row=current_row, column=1, value=f"Message: {msg.name}")
            label_cell_2 = ws.cell(row=current_row, column=2, value=f"ID: {hex(msg.frame_id)}")
            label_cell_3 = ws.cell(row=current_row, column=3, value=f"Sender(s): {senders}")
            for cell in [label_cell_1, label_cell_2, label_cell_3]:
                cell.font = header_font
                cell.fill = message_fill
                cell.border = thin_border
            current_row += 1
            current_row = write_message_table(ws, msg, current_row)

        auto_adjust_column_width(ws)

    wb.save(output_filename)
    print(f"✅ Exported to {output_filename}")

if __name__ == "__main__":
    dbc_files = [f for f in os.listdir('.') if f.endswith('.dbc')]
    if not dbc_files:
        print("❌ No .dbc files found in the current directory.")
    else:
        export_dbc_to_excel(dbc_files)
